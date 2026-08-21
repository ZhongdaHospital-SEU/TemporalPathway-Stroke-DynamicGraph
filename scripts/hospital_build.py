# -*- coding: utf-8 -*-
"""Build the 185-patient hospital-cohort harmonized dataset from the ENG template."""
from __future__ import annotations
import os, io, json
import numpy as np
import pandas as pd
from openpyxl import load_workbook

BASE = r"D:\TT paper\0811Temporal Pathway"
TPL = os.path.join(BASE, "Hospital_ICU_Stroke_Cohort_185.xlsx")
OUT = os.path.join(BASE, "work")
SCHEMA = json.load(io.open(os.path.join(BASE, "work", "harmonized_schema.json"), encoding="utf-8"))
CH = SCHEMA["channels"]                       # ['hr','rr','spo2','sbp','dbp','mbp','temp','gcs']
LABS = [l["name"] for l in SCHEMA["labs"]]    # 19 labs
LAB_STATS = SCHEMA["lab_stats"]               # n,first,last,min,max,mean
CH_STATS = SCHEMA["channel_stats"]            # first,last,min,max,mean,std,slope
DEMO = SCHEMA["demo_features"]                # age,gender_male,race_white

TS2CH = {"Heart rate (bpm)": "hr", "Resp rate (/min)": "rr", "SpO2 (%)": "spo2",
         "SBP (mmHg)": "sbp", "DBP (mmHg)": "dbp", "MAP (mmHg)": "mbp",
         "Temp (C)": "temp", "GCS Total": "gcs"}

def col(hdr, name): return hdr.index(name)

wb = load_workbook(TPL, read_only=True, data_only=True)

# ---- patient info ----
ws = wb["Patient_Info_and_Outcomes"]
rows = list(ws.iter_rows(values_only=True))
hdr = [str(x) for x in rows[0]]
data = [r for r in rows[1:] if any(x is not None and str(x).strip() != "" for x in r)]
C = {name: col(hdr, name) for name in hdr}
pid = C["Patient ID"]
info = pd.DataFrame(data, columns=hdr)
info["Patient ID"] = info["Patient ID"].astype(str)
print("[info] n =", len(info))

# ---- ICU time series ----
ws2 = wb["ICU_TimeSeries"]
rows2 = list(ws2.iter_rows(values_only=True))
hdr2 = [str(x) for x in rows2[0]]
ts = pd.DataFrame([r for r in rows2[1:] if r[0] is not None], columns=hdr2)
ts["Patient ID"] = ts["Patient ID"].astype(str)
ts = ts[ts["Patient ID"].isin(set(info["Patient ID"]))].copy()
print("[ts] rows =", len(ts), "patients =", ts["Patient ID"].nunique())

# per-channel numeric columns
ch_cols = list(TS2CH.keys())
lab_ts = {"Glucose (mmol/L)": "glucose", "Lactate (mmol/L)": "lactate"}
for c in ch_cols + list(lab_ts):
    ts[c] = pd.to_numeric(ts[c], errors="coerce")
ts["Hours since ICU admission"] = pd.to_numeric(ts["Hours since ICU admission"], errors="coerce")

def channel_stats(sub, col, ch):
    """Return dict of 7 stats for one channel from a patient's TS rows."""
    v = sub[col].dropna()
    out = {}
    if len(v) == 0:
        return out
    t = sub.loc[v.index, "Hours since ICU admission"]
    out[f"{ch}_first"] = v.iloc[0]
    out[f"{ch}_last"] = v.iloc[-1]
    out[f"{ch}_min"] = v.min()
    out[f"{ch}_max"] = v.max()
    out[f"{ch}_mean"] = v.mean()
    out[f"{ch}_std"] = v.std() if len(v) >= 2 else np.nan
    if len(v) >= 2 and v.std() > 1e-9:
        out[f"{ch}_slope"] = float(np.polyfit(t, v, 1)[0])
    else:
        out[f"{ch}_slope"] = np.nan
    return out

def lab_stats_ts(sub, lab):
    v = sub[f"Glucose (mmol/L)" if lab == "glucose" else "Lactate (mmol/L)"].dropna()
    out = {}
    if len(v) == 0:
        return out
    v_mgdl = v * 18.016 if lab == "glucose" else v
    out[f"lab_{lab}_n"] = len(v)
    out[f"lab_{lab}_first"] = v_mgdl.iloc[0]
    out[f"lab_{lab}_last"] = v_mgdl.iloc[-1]
    out[f"lab_{lab}_min"] = v_mgdl.min()
    out[f"lab_{lab}_max"] = v_mgdl.max()
    out[f"lab_{lab}_mean"] = v_mgdl.mean()
    return out

def lab_stats_single(lab, value):
    if pd.isna(value):
        return {}
    out = {f"lab_{lab}_n": 1.0, f"lab_{lab}_first": float(value),
           f"lab_{lab}_last": float(value), f"lab_{lab}_min": float(value),
           f"lab_{lab}_max": float(value), f"lab_{lab}_mean": float(value)}
    return out

seqs, masks = {}, {}
feat_rows = []
for _, p in info.iterrows():
    sid = p["Patient ID"]
    sub = ts[ts["Patient ID"] == sid].sort_values("Hours since ICU admission")
    seq = np.full((64, len(CH)), np.nan, dtype=np.float32)
    mask = np.zeros((64, len(CH)), dtype=np.float32)
    for j, (tsc, ch) in enumerate(TS2CH.items()):
        s2 = sub[["Hours since ICU admission", tsc]].dropna()
        for _, rr in s2.iterrows():
            hh = int(round(float(rr["Hours since ICU admission"])))
            if 0 <= hh < 64:
                seq[hh, j] = float(rr[tsc])
                mask[hh, j] = 1.0
    # temp C -> F
    if "temp" in TS2CH.values():
        jt = CH.index("temp")
        seq[:, jt] = seq[:, jt] * 9.0 / 5.0 + 32.0
    seqs[sid] = seq
    masks[sid] = mask

    feats = {}
    for tsc, ch in TS2CH.items():
        cs = channel_stats(sub, tsc, ch)
        if ch == "temp":  # static stats: convert C -> F to match harmonized schema
            for k in list(cs):
                if k.endswith(("_first", "_last", "_min", "_max", "_mean", "_slope")) and not pd.isna(cs[k]):
                    cs[k] = cs[k] * 9.0 / 5.0 + 32.0
        feats.update(cs)
    feats.update(lab_stats_ts(sub, "glucose"))
    feats.update(lab_stats_ts(sub, "lactate"))
    feats.update(lab_stats_single("wbc", p["First WBC"]))
    feats.update(lab_stats_single("inr", p["First INR"]))
    # demo
    feats["age"] = float(p["Age"]) if pd.notna(p["Age"]) else np.nan
    feats["gender_male"] = float(p["Sex (1M/0F)"]) if pd.notna(p["Sex (1M/0F)"]) else np.nan
    feats["race_white"] = np.nan
    feats["mrs90"] = float(p["90-day mRS"]) if pd.notna(p["90-day mRS"]) else np.nan
    feats["death"] = float(p["In-hospital death"]) if pd.notna(p["In-hospital death"]) else np.nan
    feats["lost"] = float(p["Lost to FU"]) if pd.notna(p["Lost to FU"]) else 0.0
    feats["ivt"] = float(p["IV thrombolysis"]) if pd.notna(p["IV thrombolysis"]) else 0.0
    feats["mt"] = float(p["Mech thrombectomy"]) if pd.notna(p["Mech thrombectomy"]) else 0.0
    feats["nihss"] = float(p["Admission NIHSS"]) if pd.notna(p["Admission NIHSS"]) else np.nan
    feats["gcs_admission"] = float(p["Admission GCS total"]) if pd.notna(p["Admission GCS total"]) else np.nan
    feats["age_grp_"] = p["Age"]
    feat_rows.append({"Patient ID": sid, **feats})

st = pd.DataFrame(feat_rows)
# fill any missing channel/lab columns with NaN (schema completeness check)
for ch in CH:
    for s2 in CH_STATS:
        c = f"{ch}_{s2}"
        if c not in st.columns:
            st[c] = np.nan
for lab in LABS:
    for s2 in LAB_STATS:
        c = f"lab_{lab}_{s2}"
        if c not in st.columns:
            st[c] = np.nan
for d in DEMO:
    if d not in st.columns:
        st[d] = np.nan

st.to_csv(os.path.join(OUT, "hospital_static.csv"), index=False)
np.savez(os.path.join(OUT, "hospital_seqs.npz"),
         **{f"seq_{k}": v for k, v in seqs.items()},
         **{f"mask_{k}": v for k, v in masks.items()})
print("[build] saved hospital_static.csv", st.shape)
print("[build] patients:", len(seqs), "seq shape:", seqs[list(seqs)[0]].shape)
# completeness
nch = sum(1 for c in st.columns if c in [f"{ch}_{s2}" for ch in CH for s2 in CH_STATS])
print("[build] channel-stat cols:", nch, "| non-null frac:", round(st.iloc[:, 1:].notna().mean().mean(), 3))
print("[build] mRS dist:", st["mrs90"].value_counts().sort_index().to_dict())
print("[build] good(0-2):", int((st["mrs90"] <= 2).sum()), "/", int(st["mrs90"].notna().sum()))
wb.close()

